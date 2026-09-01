#!/usr/bin/env python3
"""Build the A6 evaluator TSV from the approved marking workbook."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


HEADER = [
    "CriterionID", "Subsection", "Description", "MaxMark", "RunFrom",
    "Commands", "ExpectedResult", "Notes",
]


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").split())


def procedure_source(procedure: object, fallback: object) -> str:
    text = clean(procedure)
    match = re.match(r"Source:\s*([^.]*)\.", text, flags=re.IGNORECASE)
    if not match:
        return clean(fallback)
    source = match.group(1).strip()
    aliases = {
        "both gateways": "sh-edge-a6 and sz-edge-a6",
        "each node": (
            "sh-edge-a6, sh-user-a6, sz-edge-a6, ops-a6, services-a6, "
            "directory-a6 and network-a6"
        ),
        "both clients": "sh-user-a6 and ops-a6",
    }
    return aliases.get(source.lower(), source)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True, read_only=True)
    sheet = workbook["Expert Marking Sheet"]
    how_sheet = workbook["How to Mark"]
    safe_notes = {
        clean(row[0]): clean(row[6])
        for row in how_sheet.iter_rows(min_row=5, values_only=True)
        if row[0]
    }
    run_sources = {
        clean(row[0]): clean(row[2])
        for row in how_sheet.iter_rows(min_row=5, values_only=True)
        if row[0]
    }
    rows = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        criterion, aspect, subsection, description, expected, mark, kind, procedure = row[:8]
        if not aspect or not str(aspect).startswith(tuple("ABCDEFGHI")):
            continue
        rows.append([
            clean(aspect), clean(subsection), clean(description), clean(mark),
            procedure_source(procedure, run_sources.get(clean(aspect), clean(criterion))), clean(procedure),
            clean(expected), safe_notes.get(clean(aspect)) or "No additional note.",
        ])

    if len(rows) != 73 or abs(sum(float(row[3]) for row in rows) - 25.0) > 1e-9:
        raise SystemExit("unexpected A6 workbook content: expected 73 aspects / 25.00 marks")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
        writer.writerow(HEADER)
        writer.writerows(rows)


if __name__ == "__main__":
    main()
