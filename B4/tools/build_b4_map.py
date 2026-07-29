from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path(r"D:\Worldskills\WSTC2026\Module B\B4_marking_scheme_by_device_25_final.xlsx")
ROOT = Path(__file__).resolve().parents[1]
TARGET = ROOT / "criteria" / "b4_device_criteria_map.tsv"


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def host_key(value):
    value = clean(value)
    if value.startswith("SHA-CL01"):
        return "SHA-CL01"
    return value.upper()


wb = load_workbook(SOURCE, data_only=True)
ws = wb["Device Marking Scheme"]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[2] or not isinstance(row[7], (int, float)):
        continue
    rows.append(
        [
            host_key(row[0]),
            clean(row[1]),
            clean(row[2]),
            clean(row[3]),
            clean(row[4]),
            clean(row[5]),
            clean(row[6]),
            f"{float(row[7]):.2f}",
        ]
    )

if len(rows) != 106:
    raise SystemExit(f"Expected 106 aspects, got {len(rows)}")
if abs(sum(float(row[7]) for row in rows) - 25.0) > 0.00001:
    raise SystemExit("Mark total is not 25.00")
if len({row[2] for row in rows}) != len(rows):
    raise SystemExit("Duplicate aspect IDs")

TARGET.parent.mkdir(parents=True, exist_ok=True)
header = [
    "HostKey",
    "HostRole",
    "AspectID",
    "TaskRef",
    "Requirement",
    "VerificationCommands",
    "ExpectedResult",
    "MaxMark",
]
with TARGET.open("w", encoding="utf-8", newline="\n") as stream:
    stream.write("\t".join(header) + "\n")
    for row in rows:
        stream.write("\t".join(row) + "\n")

print(f"Wrote {len(rows)} aspects, total 25.00: {TARGET}")
