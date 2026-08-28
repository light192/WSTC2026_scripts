from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"D:\Worldskills\WSTC2026\Module B\B5_marking_scheme_corrected_final_revised.xlsx")
ROOT = Path(__file__).resolve().parents[1]
TARGET = ROOT / "criteria" / "b5_device_criteria_map.tsv"


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


wb = load_workbook(SOURCE, data_only=True)
ws = wb["Device Marking Scheme"]
rows = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0] or not isinstance(row[5], (int, float)):
        continue
    rows.append(
        [
            clean(row[1]).upper(),
            clean(row[0]),
            clean(row[2]),
            clean(row[3]),
            clean(row[4]),
            clean(row[7]),
            clean(row[8]),
            clean(row[9]),
            f"{float(row[5]):.2f}",
        ]
    )

if len(rows) != 112:
    raise SystemExit(f"Expected 112 aspects, got {len(rows)}")
if abs(sum(float(row[-1]) for row in rows) - 25.0) > 0.00001:
    raise SystemExit("Mark total is not 25.00")
if len({row[1] for row in rows}) != len(rows):
    raise SystemExit("Duplicate aspect IDs")

TARGET.parent.mkdir(parents=True, exist_ok=True)
header = [
    "HostKey", "AspectID", "TaskRef", "Category", "Requirement",
    "VerificationCommands", "ExpectedResult", "AwardGuidance", "MaxMark",
]
with TARGET.open("w", encoding="utf-8", newline="\n") as stream:
    stream.write("\t".join(header) + "\n")
    for row in rows:
        stream.write("\t".join(row) + "\n")

print(f"Wrote {len(rows)} aspects, total 25.00: {TARGET}")
